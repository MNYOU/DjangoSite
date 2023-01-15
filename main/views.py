import random
import re
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import render
from .models import Vacancy
import requests
import json


def base(request):
    # return render(request, 'main/base_template.html', {'title':'Базовая страница'})
    return render(request, 'main/base_template.html')


def home(request):
    return render(request, 'main/index.html', {"title": "Главная страница"})


def vacancies(request, vacancy_id):
    vacancies = Vacancy.objects.all()  # from database
    return HttpResponse(f"<h1>Некая вакансия с неким номером {vacancy_id}<h1>")


def relevance(request):
    path = os.path.abspath(os.path.curdir)
    book = openpyxl.load_workbook(f'{path}/main/static/main/report.xlsx')
    # book = openpyxl.load_workbook(f'{path}/DjangoSite/main/static/main/report.xlsx') # для сайта
    data = create_template(book)
    data['title'] = "Востребованность"
    return render(request, 'main/relevance.html', data)


def create_template(book):
    headings_years, data_years = get_formatted_data(book, 'Статистика по годам')
    headings_cities, data_cities = get_formatted_data(book, 'Статистика по городам', True)
    pdf_template = {'headings_years': headings_years, 'data_years': data_years,
                    'headings_cities': headings_cities, 'data_cities': data_cities}
    return pdf_template


def get_formatted_data(book, sheet_name, need_formatting=False):
    data = []
    is_heading = True
    for row in book[sheet_name]:
        if is_heading:
            naming = list(map(lambda x: x.value, row))
            is_heading = False
        else:
            row_values = list(map(lambda x: x.value, row))
            if need_formatting:
                row_values[-1] = format(row_values[-1], '.2%')
            data.append(row_values)
    return naming, data


def last_vacancies(request):
    dates = get_random_date()
    url = f'https://api.hh.ru/vacancies?text=gamedev&date_from={dates["date_from"]}&date_to={dates["date_to"]}&order_by=publication_time&only_with_salary=true'
    response = requests.get(url, headers={"User-Agent": "api-test-agent"})
    given_vacancies = response.json()['items']
    given_vacancies.reverse()
    request_method = lambda x: requests.get(f"https://api.hh.ru/vacancies/{x}",
                                            headers={"User-Agent": "api-test-agent"})
    vacancies = []
    for vacancy in given_vacancies[:5]:
        current_vacancy = request_method(vacancy['id']).json()
        current_vacancy = get_correct_vacancy(current_vacancy)
        current_vacancy['published_at'] = parse_date(current_vacancy['published_at'])
        vacancies.append(current_vacancy)
    return render(request, "main/vacancies.html",
                  {"vacancies": vacancies, 'title': 'Последнии вакансии'})


def get_random_date():
    day_from = random.randint(1, 31)
    day_to = day_from + 5
    if len(str(day_from)) == 1:
        day_from = '0' + str(day_from)
    if len(str(day_to)) == 1:
        day_to = '0' + str(day_to)
    elif int(day_to) > 26:
        day_to = '31'
    return {'date_from': f'2022-12-{day_from}', 'date_to': f'2022-12-{day_to}'}


def parse_date(s):
    if s is None or ('T' not in s and '-' not in s.split('T')):
        return ""
    time = s.split('T')[0].split('-')
    time.reverse()
    return '.'.join(time)


def get_correct_vacancy(vacancy):
    def get_correct_string(s):
        if not isinstance(s, str):
            return s
        s = re.sub(r'<[^>]*>', '', s)
        result = []
        for item in s.split('\n'):
            result.append(' '.join(item.split()))
        return '\n'.join(result)

    return {key: get_correct_string(vacancy[key]) for key in vacancy}


def geography(request):
    path = os.path.abspath(os.path.curdir)
    book = openpyxl.load_workbook(f'{path}/main/static/main/report.xlsx')
    # book = openpyxl.load_workbook(f'{path}/DjangoSite/main/static/main/report.xlsx') # для сайта
    data = create_template(book)
    data['title'] = "География"
    return render(request, 'main/geography.html', data)
